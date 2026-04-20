import base64
import importlib.util
import io
import os
import secrets
import smtplib
from smtplib import SMTPAuthenticationError
from datetime import datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, session, url_for
from openpyxl import Workbook, load_workbook
from PIL import Image

if importlib.util.find_spec("face_recognition") and importlib.util.find_spec("numpy"):
    import face_recognition
    import numpy as np
else:
    face_recognition = None
    np = None

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "attendance_secret")

BASE_DIR = Path(__file__).resolve().parent
ATTENDANCE_FILE = BASE_DIR / "attendance.xlsx"
KNOWN_FACES_DIR = BASE_DIR / "known_faces"

OTP_SENDER_EMAIL = "n4manphogat@gmail.com"
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")

ADMIN_HEADERS = ["Employee ID", "OTP Receiver Email"]
STUDENT_HEADERS = ["Roll No", "Name", "Email"]
ATTENDANCE_HEADERS = ["Roll No", "Name", "Date", "Time", "Status", "Class", "Professor", "Employee ID"]
DEFAULT_CLASS_NAME = ""
DEFAULT_PROFESSOR_NAME = ""
PRESENT_CUTOFF_TIME = os.environ.get("PRESENT_CUTOFF_TIME", "10:00")
OTP_EXPIRY_MINUTES = 5
FACE_MATCH_THRESHOLD = float(os.environ.get("FACE_MATCH_THRESHOLD", "0.55"))
RECOGNITION_FRAME_WIDTH = int(os.environ.get("RECOGNITION_FRAME_WIDTH", "640"))


def _normalize_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _find_sheet_by_headers(workbook, expected_headers: list[str]):
    expected = [_normalize_header(h) for h in expected_headers]
    for sheet in workbook.worksheets:
        first_row = [cell.value for cell in sheet[1]]
        normalized = [_normalize_header(v) for v in first_row]
        if normalized[: len(expected)] == expected:
            return sheet
    return None


def _open_workbook(data_only: bool = False):
    try:
        if ATTENDANCE_FILE.exists():
            return load_workbook(ATTENDANCE_FILE, data_only=data_only)
        return Workbook()
    except PermissionError as exc:
        raise RuntimeError(
            "Cannot open attendance.xlsx. Close the file in Excel and retry."
        ) from exc


def _save_workbook(workbook) -> None:
    try:
        workbook.save(ATTENDANCE_FILE)
    except PermissionError as exc:
        raise RuntimeError(
            "Cannot write attendance.xlsx because it is open in another app. "
            "Please close Excel and try again."
        ) from exc


def ensure_workbook() -> None:
    """Ensure attendance.xlsx contains Admin, Students and Attendance sheets."""
    wb = _open_workbook(data_only=False)
    needs_save = not ATTENDANCE_FILE.exists()

    admin_ws = (
        wb["Admin"] if "Admin" in wb.sheetnames else _find_sheet_by_headers(wb, ADMIN_HEADERS)
    )
    students_ws = (
        wb["Students"]
        if "Students" in wb.sheetnames
        else _find_sheet_by_headers(wb, ["Roll No", "Name"])
    )
    attendance_ws = (
        wb["Attendance"] if "Attendance" in wb.sheetnames else _find_sheet_by_headers(wb, ATTENDANCE_HEADERS)
    )

    if admin_ws is None:
        admin_ws = wb.create_sheet("Admin")
        admin_ws.append(ADMIN_HEADERS)
        admin_ws.append(["EMP001", "anyone@example.com"])
        needs_save = True
    elif admin_ws.title != "Admin":
        admin_ws.title = "Admin"
        needs_save = True

    if students_ws is None:
        students_ws = wb.create_sheet("Students")
        students_ws.append(STUDENT_HEADERS)
        needs_save = True
    else:
        if students_ws.title != "Students":
            students_ws.title = "Students"
            needs_save = True
        header = [_normalize_header(v) for v in [cell.value for cell in students_ws[1]]]
        if len(header) < 3:
            students_ws.cell(row=1, column=3, value="Email")
            needs_save = True
        elif header[2] in {"password", "pass", "pwd"}:
            students_ws.cell(row=1, column=3, value="Email")
            needs_save = True

    if attendance_ws is None:
        attendance_ws = wb.create_sheet("Attendance")
        attendance_ws.append(ATTENDANCE_HEADERS)
        needs_save = True
    else:
        if attendance_ws.title != "Attendance":
            attendance_ws.title = "Attendance"
            needs_save = True
        attendance_header = [_normalize_header(v) for v in [cell.value for cell in attendance_ws[1]]]
        if len(attendance_header) < 6:
            attendance_ws.cell(row=1, column=6, value="Class")
            needs_save = True
        if len(attendance_header) < 7:
            attendance_ws.cell(row=1, column=7, value="Professor")
            needs_save = True
        if len(attendance_header) < 8:
            attendance_ws.cell(row=1, column=8, value="Employee ID")
            needs_save = True

    # Seed workbook rows to match default template when sheets are empty.
    if students_ws.max_row <= 1:
        students_ws.append([101, "Student One", "student101@example.com"])
        students_ws.append([102, "Student Two", "student102@example.com"])
        needs_save = True

    attendance_data_rows = [
        row for row in attendance_ws.iter_rows(min_row=2, values_only=True)
        if row and any(cell not in (None, "") for cell in row)
    ]
    if not attendance_data_rows:
        attendance_ws.append([101, "Student One", "14-Feb", "09:50", "PRESENT", "", "", "EMP001"])
        attendance_ws.append([101, "Student One", "14-Feb", "10:40", "ABSENT", "", "", "EMP001"])
        needs_save = True

    if wb.active.title not in {"Admin", "Students", "Attendance"}:
        wb.active = wb["Admin"]
        needs_save = True

    if needs_save:
        _save_workbook(wb)


def read_students() -> dict[str, dict[str, str]]:
    ensure_workbook()
    wb = _open_workbook(data_only=True)
    ws = wb["Students"]

    students: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        roll = str(row[0]).strip()
        students[roll] = {
            "name": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            "email": str(row[2]).strip() if len(row) > 2 and row[2] else "",
        }
    return students


def read_admin_settings() -> dict[str, str]:
    ensure_workbook()
    wb = _open_workbook(data_only=True)
    ws = wb["Admin"]

    employee_id = "EMP001"
    otp_receiver = OTP_SENDER_EMAIL
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        if row[0]:
            employee_id = str(row[0]).strip()
        if len(row) > 1 and row[1]:
            otp_receiver = str(row[1]).strip()
            if otp_receiver.lower() == "n4manphogat.gmail.com":
                otp_receiver = "anyone@example.com"
        break

    return {"employee_id": employee_id, "otp_receiver": otp_receiver}


def _format_excel_date(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    # Handles strings like "2026-01-01 00:00:00"
    if " " in text and text.endswith("00:00:00"):
        return text.split(" ", 1)[0]
    return text


def _format_excel_time(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    return str(value).strip()

def get_attendance_for_roll(roll: str) -> list[dict[str, str]]:
    ensure_workbook()
    wb = _open_workbook(data_only=True)
    ws = wb["Attendance"]

    records: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or str(row[0]).strip() != str(roll):
            continue
        records.append({
            "date": _format_excel_date(row[2]),
            "time": _format_excel_time(row[3]),
            "status": str(row[4]).strip() if row[4] is not None else "",
            "class_name": str(row[5]).strip() if len(row) > 5 and row[5] else DEFAULT_CLASS_NAME,
            "professor_name": str(row[6]).strip() if len(row) > 6 and row[6] else DEFAULT_PROFESSOR_NAME,
            "employee_id": str(row[7]).strip() if len(row) > 7 and row[7] else "",
        })
    return records


def _attendance_status_for_now(now: datetime) -> str:
    cutoff_raw = PRESENT_CUTOFF_TIME.strip()
    try:
        cutoff_time = datetime.strptime(cutoff_raw, "%H:%M").time()
    except ValueError:
        cutoff_time = datetime.strptime("10:00", "%H:%M").time()
    return "PRESENT" if now.time() <= cutoff_time else "ABSENT"


def mark_attendance(roll: str, name: str, employee_id: str) -> dict[str, str]:
    ensure_workbook()
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Always mark PRESENT when a student is successfully recognized by the system.
    status = "PRESENT"

    wb = _open_workbook(data_only=False)
    ws = wb["Attendance"]
    ws.append([roll, name, date_str, time_str, status, DEFAULT_CLASS_NAME, DEFAULT_PROFESSOR_NAME, employee_id])
    _save_workbook(wb)

    return {"roll": roll, "name": name, "date": date_str, "time": time_str, "status": status, "employee_id": employee_id}


def _smtp_password() -> str:
    return GMAIL_APP_PASSWORD.replace(" ", "").strip()


def _is_smtp_configured() -> bool:
    return bool(_smtp_password())


def _is_smtp_error_recoverable(error: Exception) -> bool:
    text = str(error).lower()
    return (
        "gmail authentication failed" in text
        or "set gmail_app_password" in text
        or "gmail_app_password is empty" in text
    )

def _send_otp(otp: str, receiver_email: str, audience: str) -> None:
    if not GMAIL_APP_PASSWORD:
        raise RuntimeError("Set GMAIL_APP_PASSWORD env var to send OTP email.")

    smtp_password = _smtp_password()
    if not smtp_password:
        raise RuntimeError("GMAIL_APP_PASSWORD is empty after trimming spaces.")

    message = EmailMessage()
    message["Subject"] = f"{audience} Login OTP"
    message["From"] = OTP_SENDER_EMAIL
    message["To"] = receiver_email
    message.set_content(f"Your OTP is: {otp}\nValid for {OTP_EXPIRY_MINUTES} minutes.\n")

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(OTP_SENDER_EMAIL, smtp_password)
            smtp.send_message(message)
    except SMTPAuthenticationError as exc:
        raise RuntimeError(
            "Gmail authentication failed (535). Use a Gmail App Password for "
            "n4manphogat@gmail.com with 2-Step Verification enabled, and set it "
            "as GMAIL_APP_PASSWORD (you can paste with spaces; app strips them)."
        ) from exc


def admin_required(route_func):
    @wraps(route_func)
    def wrapper(*args, **kwargs):
        if session.get("admin_authenticated") is not True:
            return redirect(url_for("admin_login"))
        return route_func(*args, **kwargs)

    return wrapper


def student_required(route_func):
    @wraps(route_func)
    def wrapper(*args, **kwargs):
        if "student_roll" not in session:
            return redirect(url_for("student_login"))
        return route_func(*args, **kwargs)

    return wrapper


def _decode_data_url_to_bytes(data_url: str) -> bytes:
    _, encoded = data_url.split(",", 1) if "," in data_url else ("", data_url)
    if not encoded:
        raise ValueError("Empty image payload")
    return base64.b64decode(encoded)


def _load_known_face_encodings() -> list[dict[str, object]]:
    if face_recognition is None:
        return []

    students = read_students()
    known_faces: list[dict[str, object]] = []

    if not KNOWN_FACES_DIR.exists():
        return known_faces

    for roll, metadata in students.items():
        folder = KNOWN_FACES_DIR / roll
        if not folder.exists() or not folder.is_dir():
            continue

        for image_file in folder.iterdir():
            if image_file.suffix.lower() not in {".jpg", ".jpeg", ".png"}:
                continue
            image = face_recognition.load_image_file(str(image_file))
            encodings = face_recognition.face_encodings(image)
            if encodings:
                known_faces.append({"roll": roll, "name": metadata["name"], "encoding": encodings[0]})
    return known_faces


def _resize_frame_for_recognition(image: Image.Image) -> Image.Image:
    if image.width <= RECOGNITION_FRAME_WIDTH:
        return image
    ratio = RECOGNITION_FRAME_WIDTH / float(image.width)
    new_size = (RECOGNITION_FRAME_WIDTH, max(1, int(image.height * ratio)))
    return image.resize(new_size, Image.Resampling.LANCZOS)

def _recognize_student_from_frame(data_url: str) -> dict[str, str] | None:
    if face_recognition is None or np is None:
        raise RuntimeError("face_recognition dependency is missing. Install with: pip install -r requirements-face.txt")

    payload = _decode_data_url_to_bytes(data_url)
    pil_image = Image.open(io.BytesIO(payload)).convert("RGB")
    pil_image = _resize_frame_for_recognition(pil_image)
    frame_encodings = face_recognition.face_encodings(np.array(pil_image))
    if not frame_encodings:
        return None

    known_faces = _load_known_face_encodings()
    if not known_faces:
        return None

    known_encodings = [known["encoding"] for known in known_faces]
    for frame_encoding in frame_encodings:
        distances = face_recognition.face_distance(known_encodings, frame_encoding)
        if len(distances) == 0:
            continue
        best_idx = int(distances.argmin())
        if distances[best_idx] <= FACE_MATCH_THRESHOLD:
            best_match = known_faces[best_idx]
            return {"roll": str(best_match["roll"]), "name": str(best_match["name"])}
    return None


@app.errorhandler(RuntimeError)
def handle_runtime_error(error):
    message = str(error)
    if request.path.startswith("/admin/recognize"):
        return jsonify({"success": False, "message": message}), 500
    if request.path.startswith("/student/login"):
        return render_template("student_login.html", error=message), 500
    if request.path.startswith("/admin/login"):
        return render_template("admin_login.html", error=message), 500
    return render_template("index.html", error=message), 500

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/student/login", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        action = request.form.get("action", "login")
        roll = request.form.get("roll", "").strip()
        students = read_students()
        student = students.get(roll)

        if not student:
            return render_template("student_login.html", error="Invalid roll number", roll=roll)

        student_email = student.get("email", "")
        if "@" not in student_email:
            return render_template(
                "student_login.html",
                error="Student email missing in Excel. Update Students sheet column 3.",
                roll=roll,
            )

        if action == "send_otp":
            otp = f"{secrets.randbelow(1_000_000):06d}"
            expiry = datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES)
            session["pending_student_roll"] = roll
            session["student_otp"] = otp
            session["student_otp_expiry"] = expiry.isoformat()
            try:
                _send_otp(otp, student_email, "Student")
            except Exception as exc:
                return render_template("student_login.html", error=f"OTP sending failed: {exc}", roll=roll)

            return render_template(
                "student_login.html",
                message=f"OTP sent from {OTP_SENDER_EMAIL} to {student_email}.",
                roll=roll,
            )

        entered_otp = request.form.get("otp", "").strip()
        saved_otp = session.get("student_otp")
        saved_roll = session.get("pending_student_roll")
        expiry_str = session.get("student_otp_expiry")

        if not entered_otp or not saved_otp or not expiry_str:
            return render_template("student_login.html", error="Please request OTP first.", roll=roll)

        if saved_roll != roll:
            return render_template("student_login.html", error="OTP belongs to different roll number.", roll=roll)

        if datetime.now() > datetime.fromisoformat(expiry_str):
            return render_template("student_login.html", error="OTP expired. Please request a new OTP.", roll=roll)

        if entered_otp != saved_otp:
            return render_template("student_login.html", error="Invalid OTP", roll=roll)

        session.clear()
        session["student_roll"] = roll
        return redirect(url_for("student_dashboard"))

    return render_template("student_login.html")


@app.route("/student/dashboard")
@student_required
def student_dashboard():
    roll = session["student_roll"]
    records = get_attendance_for_roll(roll)
    return render_template("student_dashboard.html", roll=roll, records=records)


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        action = request.form.get("action", "login")
        employee_id = request.form.get("employee_id", "").strip()

        admin_settings = read_admin_settings()
        admin_employee_id = admin_settings["employee_id"]
        otp_receiver = admin_settings["otp_receiver"]

        if employee_id != admin_employee_id:
            return render_template("admin_login.html", error="Invalid employee ID", employee_id=employee_id)

        if action == "send_otp":
            otp = f"{secrets.randbelow(1_000_000):06d}"
            expiry = datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES)
            session["pending_admin_employee_id"] = employee_id
            session["admin_otp"] = otp
            session["admin_otp_expiry"] = expiry.isoformat()

            if _is_smtp_configured():
                try:
                    _send_otp(otp, otp_receiver, "Admin")
                    return render_template(
                        "admin_login.html",
                        message=f"OTP sent from {OTP_SENDER_EMAIL} to {otp_receiver}.",
                        employee_id=employee_id,
                    )
                except Exception as exc:
                    if _is_smtp_error_recoverable(exc):
                        return render_template(
                            "admin_login.html",
                            message=(
                                "Email OTP failed due to Gmail password/auth config. "
                                f"Use this OTP to continue login: {otp}"
                            ),
                            error=f"Email send skipped: {exc}",
                            employee_id=employee_id,
                        )
                    return render_template("admin_login.html", error=f"OTP sending failed: {exc}", employee_id=employee_id)

            return render_template(
                "admin_login.html",
                message=(
                    "Email OTP is not configured (missing GMAIL_APP_PASSWORD). "
                    f"Use this OTP to continue login: {otp}"
                ),
                employee_id=employee_id,
            )

        entered_otp = request.form.get("otp", "").strip()
        saved_otp = session.get("admin_otp")
        saved_employee_id = session.get("pending_admin_employee_id")
        expiry_str = session.get("admin_otp_expiry")

        if not entered_otp or not saved_otp or not expiry_str:
            return render_template("admin_login.html", error="Please request OTP first.", employee_id=employee_id)

        if saved_employee_id != employee_id:
            return render_template("admin_login.html", error="OTP was generated for different employee ID.", employee_id=employee_id)

        if datetime.now() > datetime.fromisoformat(expiry_str):
            return render_template("admin_login.html", error="OTP expired. Please request a new OTP.", employee_id=employee_id)

        if entered_otp != saved_otp:
            return render_template("admin_login.html", error="Invalid OTP", employee_id=employee_id)

        session.clear()
        session["admin_authenticated"] = True
        session["admin_employee_id"] = employee_id
        return redirect(url_for("admin_camera"))

    return render_template("admin_login.html")


@app.route("/admin/camera")
@admin_required
def admin_camera():
    return render_template("admin_camera.html")


@app.route("/admin/recognize", methods=["POST"])
@admin_required
def admin_recognize():
    payload = request.get_json(silent=True) or {}
    image_data = payload.get("image")

    if not image_data:
        return jsonify({"success": False, "message": "Image data is required"}), 400

    try:
        recognized = _recognize_student_from_frame(image_data)
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception:
        return jsonify({"success": False, "message": "Unable to process camera frame."}), 400

    if not recognized:
        return jsonify({"success": False, "message": "No known student face matched."})

    employee_id = str(session.get("admin_employee_id", "")).strip() or read_admin_settings()["employee_id"]
    entry = mark_attendance(recognized["roll"], recognized["name"], employee_id)
    return jsonify({"success": True, "message": "Attendance marked successfully.", "attendance": entry})


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


if __name__ == "__main__":
    ensure_workbook()
    app.run(debug=True)